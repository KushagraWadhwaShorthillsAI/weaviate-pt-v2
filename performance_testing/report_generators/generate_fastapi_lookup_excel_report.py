"""
Generate Excel performance test report for FastAPI lookup tests.
Creates a clean, organized Excel file with formatted tables comparing BM25 vs Hybrid,
with Sync vs Async bifurcation for each metric.
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))

import os
import csv
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Missing required packages. Install with:")
    print("   pip install openpyxl")
    sys.exit(1)


def parse_stats_csv(filepath):
    """Parse Locust stats CSV file"""
    stats = []
    try:
        with open(filepath, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                stats.append(row)
        return stats
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return None


def extract_key_metrics(stats):
    """Extract key metrics from stats"""
    if not stats:
        return None
    
    # Get aggregated row (last row or row with "Aggregated" name)
    aggregated = None
    for row in stats:
        if row.get('Name') == 'Aggregated' or row.get('Type') == 'Aggregated':
            aggregated = row
            break
    
    if not aggregated:
        aggregated = stats[0] if stats else None
    
    if not aggregated:
        return None
    
    try:
        return {
            'total_requests': int(aggregated.get('Request Count', 0)),
            'failures': int(aggregated.get('Failure Count', 0)),
            'avg_response': float(aggregated.get('Average Response Time', 0)),
            'min_response': float(aggregated.get('Min Response Time', 0)),
            'max_response': float(aggregated.get('Max Response Time', 0)),
            'median_response': float(aggregated.get('Median Response Time', 0)),
            'p95_response': float(aggregated.get('95%', 0)),
            'p99_response': float(aggregated.get('99%', 0)),
            'rps': float(aggregated.get('Requests/s', 0)),
            'failure_rate': (int(aggregated.get('Failure Count', 0)) / max(int(aggregated.get('Request Count', 1)), 1)) * 100
        }
    except Exception as e:
        print(f"Error extracting metrics: {e}")
        return None


def scan_fastapi_lookup_reports():
    """Scan all lookup_* folders and gather data"""
    results = defaultdict(lambda: defaultdict(dict))
    
    # Get RF, Limit, and User Counts from environment variables (set by run script)
    rf_value = os.environ.get('PT_RF_VALUE', 'current')
    limit = os.environ.get('PT_LIMIT', '200')
    user_counts_str = os.environ.get('PT_USER_COUNTS', '')
    
    # Parse user counts from environment variable (space-separated string)
    if user_counts_str:
        user_counts = [uc.strip() for uc in user_counts_str.split() if uc.strip()]
    else:
        # Fallback: try to detect from folders if not provided
        import glob
        base_dir = '../reports/multi_collection'
        # Try new simplified pattern first: lookup_RF{rf}_U{users}_L{limit}
        pattern_new = os.path.join(base_dir, f"lookup_RF{rf_value}_U*_L{limit}")
        # Fallback to old pattern: fastapi_lookup_RF{rf}_Users{users}_Limit{limit}
        pattern_old = os.path.join(base_dir, f"fastapi_lookup_RF{rf_value}_Users*_Limit{limit}")
        folders = glob.glob(pattern_new) + glob.glob(pattern_old)
        user_counts = []
        for folder in folders:
            folder_name = os.path.basename(folder)
            try:
                # Try new pattern: lookup_RF{rf}_U{users}_L{limit}
                if folder_name.startswith('lookup_RF'):
                    parts = folder_name.split('_')
                    for i, part in enumerate(parts):
                        if part.startswith('U') and len(part) > 1:
                            user_count = part[1:]  # Remove 'U' prefix
                            if user_count.isdigit() and user_count not in user_counts:
                                user_counts.append(user_count)
                            break
                # Try old pattern: fastapi_lookup_RF{rf}_Users{users}_Limit{limit}
                else:
                    parts = folder_name.split('_')
                    for i, part in enumerate(parts):
                        if part == 'Users' and i + 1 < len(parts):
                            user_count = parts[i + 1]
                            if user_count not in user_counts:
                                user_counts.append(user_count)
                            break
            except Exception:
                continue
    
    if not user_counts:
        print(f"⚠️  No user counts found (checked PT_USER_COUNTS env var and folder scan)")
        return results
    
    print(f"📂 Processing user counts: {', '.join(sorted(user_counts, key=lambda x: int(x)))}")
    
    # FastAPI lookup reports folders
    base_dir = '../reports/multi_collection'
    
    # Search types for FastAPI lookup (BM25 and Hybrid)
    search_types = ['graphql_lookup_sync_bm25', 'graphql_lookup_async_bm25', 'graphql_lookup_sync', 'graphql_lookup_async']
    
    for user_count in sorted(user_counts, key=lambda x: int(x)):
        # Try new simplified pattern first: lookup_RF{rf}_U{users}_L{limit}
        folder_new = os.path.join(base_dir, f"lookup_RF{rf_value}_U{user_count}_L{limit}")
        # Fallback to old pattern: fastapi_lookup_RF{rf}_Users{users}_Limit{limit}
        folder_old = os.path.join(base_dir, f"fastapi_lookup_RF{rf_value}_Users{user_count}_Limit{limit}")
        folder = folder_new if os.path.exists(folder_new) else folder_old
        
        if not os.path.exists(folder):
            print(f"⚠️  Folder not found: {folder}")
            continue
        
        for search_type in search_types:
            stats_file = os.path.join(folder, f"{search_type}_stats.csv")
            
            if os.path.exists(stats_file):
                stats = parse_stats_csv(stats_file)
                metrics = extract_key_metrics(stats)
                
                # Only add if metrics exist and are non-zero
                if metrics and metrics.get('total_requests', 0) > 0:
                    results[user_count][search_type] = metrics
                    print(f"✓ Loaded: {folder}/{search_type}_stats.csv")
                elif metrics:
                    # File exists but has zero values, skip silently
                    pass
                else:
                    print(f"⚠️  Could not extract metrics from: {stats_file}")
            else:
                print(f"⚠️  File not found: {stats_file}")
    
    return results


def create_excel_report(results, rf_value='3', limit='200'):
    """Create formatted Excel report with Sync/Async bifurcation"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Comparison"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_header_font = Font(bold=True, size=12)
    latency_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    latency_font = Font(color="CC0000", bold=True)  # Red text for latency
    throughput_fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    throughput_font = Font(color="006600", bold=True)  # Green text for throughput
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:D{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"FastAPI Lookup Performance Test Report - RF {rf_value} | Limit {limit}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 2
    
    # Subtitle
    ws.merge_cells(f'A{current_row}:D{current_row}')
    subtitle_cell = ws[f'A{current_row}']
    subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 3
    
    # Process each user count
    user_counts = sorted(results.keys(), key=lambda x: int(x))
    
    for user_count in user_counts:
        # Section header
        ws.merge_cells(f'A{current_row}:D{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = f"{user_count} users"
        section_cell.fill = section_header_fill
        section_cell.font = section_header_font
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Table header - Method | Metric | Sync | Async
        headers = ['Method', 'Metric', 'Sync', 'Async']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        current_row += 1
        
        # BM25 data
        bm25_sync_metrics = results[user_count].get('graphql_lookup_sync_bm25', {})
        bm25_async_metrics = results[user_count].get('graphql_lookup_async_bm25', {})
        
        if bm25_sync_metrics or bm25_async_metrics:
            # BM25 Latency row
            ws.cell(row=current_row, column=1, value='BM25').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Latency (ms)').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            
            # Sync latency
            sync_latency = bm25_sync_metrics.get('avg_response', 0) if bm25_sync_metrics else 0
            sync_latency_cell = ws.cell(row=current_row, column=3, value=round(sync_latency, 1) if sync_latency > 0 else '-')
            sync_latency_cell.fill = latency_fill
            sync_latency_cell.font = latency_font
            sync_latency_cell.border = border
            sync_latency_cell.alignment = center_align
            
            # Async latency
            async_latency = bm25_async_metrics.get('avg_response', 0) if bm25_async_metrics else 0
            async_latency_cell = ws.cell(row=current_row, column=4, value=round(async_latency, 1) if async_latency > 0 else '-')
            async_latency_cell.fill = latency_fill
            async_latency_cell.font = latency_font
            async_latency_cell.border = border
            async_latency_cell.alignment = center_align
            
            current_row += 1
            
            # BM25 Throughput row
            ws.cell(row=current_row, column=1, value='BM25').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Throughput (req/s)').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            
            # Sync throughput
            sync_throughput = bm25_sync_metrics.get('rps', 0) if bm25_sync_metrics else 0
            sync_throughput_cell = ws.cell(row=current_row, column=3, value=round(sync_throughput, 2) if sync_throughput > 0 else '-')
            sync_throughput_cell.fill = throughput_fill
            sync_throughput_cell.font = throughput_font
            sync_throughput_cell.border = border
            sync_throughput_cell.alignment = center_align
            
            # Async throughput
            async_throughput = bm25_async_metrics.get('rps', 0) if bm25_async_metrics else 0
            async_throughput_cell = ws.cell(row=current_row, column=4, value=round(async_throughput, 2) if async_throughput > 0 else '-')
            async_throughput_cell.fill = throughput_fill
            async_throughput_cell.font = throughput_font
            async_throughput_cell.border = border
            async_throughput_cell.alignment = center_align
            
            current_row += 1
        
        # Hybrid data
        hybrid_sync_metrics = results[user_count].get('graphql_lookup_sync', {})
        hybrid_async_metrics = results[user_count].get('graphql_lookup_async', {})
        
        if hybrid_sync_metrics or hybrid_async_metrics:
            # Hybrid Latency row
            ws.cell(row=current_row, column=1, value='Hybrid0.9').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Latency (ms)').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            
            # Sync latency
            sync_latency = hybrid_sync_metrics.get('avg_response', 0) if hybrid_sync_metrics else 0
            sync_latency_cell = ws.cell(row=current_row, column=3, value=round(sync_latency, 1) if sync_latency > 0 else '-')
            sync_latency_cell.fill = latency_fill
            sync_latency_cell.font = latency_font
            sync_latency_cell.border = border
            sync_latency_cell.alignment = center_align
            
            # Async latency
            async_latency = hybrid_async_metrics.get('avg_response', 0) if hybrid_async_metrics else 0
            async_latency_cell = ws.cell(row=current_row, column=4, value=round(async_latency, 1) if async_latency > 0 else '-')
            async_latency_cell.fill = latency_fill
            async_latency_cell.font = latency_font
            async_latency_cell.border = border
            async_latency_cell.alignment = center_align
            
            current_row += 1
            
            # Hybrid Throughput row
            ws.cell(row=current_row, column=1, value='Hybrid0.9').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Throughput (req/s)').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            
            # Sync throughput
            sync_throughput = hybrid_sync_metrics.get('rps', 0) if hybrid_sync_metrics else 0
            sync_throughput_cell = ws.cell(row=current_row, column=3, value=round(sync_throughput, 2) if sync_throughput > 0 else '-')
            sync_throughput_cell.fill = throughput_fill
            sync_throughput_cell.font = throughput_font
            sync_throughput_cell.border = border
            sync_throughput_cell.alignment = center_align
            
            # Async throughput
            async_throughput = hybrid_async_metrics.get('rps', 0) if hybrid_async_metrics else 0
            async_throughput_cell = ws.cell(row=current_row, column=4, value=round(async_throughput, 2) if async_throughput > 0 else '-')
            async_throughput_cell.fill = throughput_fill
            async_throughput_cell.font = throughput_font
            async_throughput_cell.border = border
            async_throughput_cell.alignment = center_align
            
            current_row += 1
        
        # Add spacing between sections
        current_row += 2
    
    # Adjust column widths for better formatting
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    return wb


def main():
    """Main function"""
    print("=" * 70)
    print("FASTAPI LOOKUP EXCEL PERFORMANCE REPORT GENERATOR")
    print("=" * 70)
    print("\nScanning FastAPI lookup reports folders...")
    print("-" * 70)
    
    # Scan all reports
    results = scan_fastapi_lookup_reports()
    
    if not results:
        print("\n❌ No data found!")
        print("   Make sure you have lookup_* folders with CSV files")
        print("   Expected pattern: lookup_RF{rf}_U{count}_L{limit}/")
        print("   (or old pattern: fastapi_lookup_RF{rf}_Users{count}_Limit{limit}/)")
        return 1
    
    print("\n" + "-" * 70)
    print(f"✅ Found data for {len(results)} user counts")
    print(f"   User counts: {', '.join(sorted(results.keys(), key=lambda x: int(x)))}")
    
    # Get RF and Limit from environment variables
    rf_value = os.environ.get('PT_RF_VALUE', 'current')
    limit = os.environ.get('PT_LIMIT', '200')
    
    # Generate Excel report
    print("\n📝 Generating Excel report...")
    wb = create_excel_report(results, rf_value=rf_value, limit=limit)
    
    # Simplified naming: lookup_combined_RF{rf}_U{users}_L{limit}.xlsx
    user_counts_sorted = sorted(results.keys(), key=lambda x: int(x))
    if len(user_counts_sorted) == 1:
        users_str = f"U{user_counts_sorted[0]}"
    else:
        # Multiple user counts: U100-200-300
        users_str = f"U{'-'.join(user_counts_sorted)}"
    
    output_file = f"../reports/multi_collection/lookup_combined_RF{rf_value}_{users_str}_L{limit}.xlsx"
    
    wb.save(output_file)
    
    print(f"✅ Created: {output_file}")
    print(f"   Location: {os.path.abspath(output_file)}")
    print("\n" + "=" * 70)
    print("EXCEL REPORT GENERATED!")
    print("=" * 70)
    print(f"\nOpen the report:")
    print(f"   {output_file}")
    print("=" * 70)
    
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())

