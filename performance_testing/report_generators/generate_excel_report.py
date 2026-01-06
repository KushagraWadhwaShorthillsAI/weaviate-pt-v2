"""
Generate Excel performance test report from multiple test runs.
Creates a clean, organized Excel file with formatted tables comparing results across user counts.
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))

import os
import csv
from collections import defaultdict
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Missing required packages. Install with:")
    print("   pip install pandas openpyxl")
    sys.exit(1)


def parse_stats_csv(filepath):
    """Parse Locust stats CSV file"""
    stats = []
    try:
        with open(filepath, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                stats.append(row)
        return stats
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return None


def extract_key_metrics(stats):
    """Extract key metrics from stats"""
    if not stats:
        return None
    
    # Get aggregated row (last row or row with "Aggregated" name)
    aggregated = None
    for row in stats:
        if row.get('Name') == 'Aggregated' or row.get('Type') == 'Aggregated':
            aggregated = row
            break
    
    if not aggregated:
        aggregated = stats[0] if stats else None
    
    if not aggregated:
        return None
    
    try:
        return {
            'total_requests': int(aggregated.get('Request Count', 0)),
            'failures': int(aggregated.get('Failure Count', 0)),
            'avg_response': float(aggregated.get('Average Response Time', 0)),
            'min_response': float(aggregated.get('Min Response Time', 0)),
            'max_response': float(aggregated.get('Max Response Time', 0)),
            'median_response': float(aggregated.get('Median Response Time', 0)),
            'p95_response': float(aggregated.get('95%', 0)),
            'p99_response': float(aggregated.get('99%', 0)),
            'rps': float(aggregated.get('Requests/s', 0)),
            'failure_rate': (int(aggregated.get('Failure Count', 0)) / max(int(aggregated.get('Request Count', 1)), 1)) * 100
        }
    except Exception as e:
        print(f"Error extracting metrics: {e}")
        return None


def scan_all_user_reports():
    """Scan reports for all user counts and gather data"""
    results = defaultdict(lambda: defaultdict(dict))
    
    base_dir = '../reports/multi_collection'
    rf_value = os.environ.get('PT_RF_VALUE', '3')
    user_counts = ['100', '200', '300']
    limits = ['200']  # Focus on limit 200 for main comparison
    search_types = ['bm25', 'hybrid_09']  # Focus on BM25 and Hybrid0.9
    
    for user_count in user_counts:
        for limit in limits:
            folder = os.path.join(base_dir, f"reports_RF{rf_value}_Users{user_count}_Limit{limit}")
            
            if not os.path.exists(folder):
                print(f"⚠️  Folder not found: {folder}")
                continue
            
            for search_type in search_types:
                stats_file = os.path.join(folder, f"{search_type}_stats.csv")
                
                if os.path.exists(stats_file):
                    stats = parse_stats_csv(stats_file)
                    metrics = extract_key_metrics(stats)
                    
                    if metrics and metrics.get('total_requests', 0) > 0:
                        results[user_count][search_type] = metrics
                        print(f"✓ Loaded: {folder}/{search_type}_stats.csv")
    
    return results


def create_excel_report(results, rf_value='3'):
    """Create formatted Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Report"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_header_font = Font(bold=True, size=12)
    latency_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    throughput_fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:E{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"Weaviate Performance Test Report - RF {rf_value}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 2
    
    # Subtitle
    ws.merge_cells(f'A{current_row}:E{current_row}')
    subtitle_cell = ws[f'A{current_row}']
    subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 3
    
    # Process each user count
    user_counts = sorted(results.keys(), key=lambda x: int(x))
    
    for user_count in user_counts:
        # Section header
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = f"{user_count} users"
        section_cell.fill = section_header_fill
        section_cell.font = section_header_font
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Table header
        header_row = current_row
        headers = ['Method', 'Metric', 'Value']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        current_row += 1
        
        # BM25 data
        if 'bm25' in results[user_count]:
            bm25_metrics = results[user_count]['bm25']
            
            # BM25 Latency
            ws.cell(row=current_row, column=1, value='BM25').border = border
            ws.cell(row=current_row, column=2, value='Latency (ms)').border = border
            latency_cell = ws.cell(row=current_row, column=3, value=round(bm25_metrics['avg_response'], 1))
            latency_cell.fill = latency_fill
            latency_cell.border = border
            latency_cell.alignment = center_align
            current_row += 1
            
            # BM25 Throughput
            ws.cell(row=current_row, column=1, value='BM25').border = border
            ws.cell(row=current_row, column=2, value='Throughput (req/s)').border = border
            throughput_cell = ws.cell(row=current_row, column=3, value=round(bm25_metrics['rps'], 2))
            throughput_cell.fill = throughput_fill
            throughput_cell.border = border
            throughput_cell.alignment = center_align
            current_row += 1
        
        # Hybrid0.9 data
        if 'hybrid_09' in results[user_count]:
            hybrid_metrics = results[user_count]['hybrid_09']
            
            # Hybrid0.9 Latency
            ws.cell(row=current_row, column=1, value='Hybrid0.9').border = border
            ws.cell(row=current_row, column=2, value='Latency (ms)').border = border
            latency_cell = ws.cell(row=current_row, column=3, value=round(hybrid_metrics['avg_response'], 1))
            latency_cell.fill = latency_fill
            latency_cell.border = border
            latency_cell.alignment = center_align
            current_row += 1
            
            # Hybrid0.9 Throughput
            ws.cell(row=current_row, column=1, value='Hybrid0.9').border = border
            ws.cell(row=current_row, column=2, value='Throughput (req/s)').border = border
            throughput_cell = ws.cell(row=current_row, column=3, value=round(hybrid_metrics['rps'], 2))
            throughput_cell.fill = throughput_fill
            throughput_cell.border = border
            throughput_cell.alignment = center_align
            current_row += 1
        
        # Add spacing between sections
        current_row += 2
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    
    return wb


def create_comparison_excel_report(all_results, rf_value='3'):
    """Create comparison Excel report with Previous vs New format (matching image structure)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Comparison"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_header_font = Font(bold=True, size=12)
    latency_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    latency_font = Font(color="CC0000", bold=True)  # Red text for latency
    throughput_fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    throughput_font = Font(color="006600", bold=True)  # Green text for throughput
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:C{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"Weaviate Performance Test Report - RF {rf_value}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 2
    
    # Subtitle
    ws.merge_cells(f'A{current_row}:C{current_row}')
    subtitle_cell = ws[f'A{current_row}']
    subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 3
    
    # Process each user count
    user_counts = sorted(all_results.keys(), key=lambda x: int(x))
    
    for user_count in user_counts:
        # Section header
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = f"{user_count} users"
        section_cell.fill = section_header_fill
        section_cell.font = section_header_font
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Table header - simplified format with single value column
        headers = ['Method', 'Metric', 'Value']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        current_row += 1
        
        # BM25 data
        if 'bm25' in all_results[user_count]:
            bm25_metrics = all_results[user_count]['bm25']
            
            # BM25 Latency row
            ws.cell(row=current_row, column=1, value='BM25').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Latency').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            latency_cell = ws.cell(row=current_row, column=3, value=round(bm25_metrics['avg_response'], 1))
            latency_cell.fill = latency_fill
            latency_cell.font = latency_font
            latency_cell.border = border
            latency_cell.alignment = center_align
            current_row += 1
            
            # BM25 Throughput row
            ws.cell(row=current_row, column=1, value='BM25').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Throughput').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            throughput_cell = ws.cell(row=current_row, column=3, value=round(bm25_metrics['rps'], 2))
            throughput_cell.fill = throughput_fill
            throughput_cell.font = throughput_font
            throughput_cell.border = border
            throughput_cell.alignment = center_align
            current_row += 1
        
        # Hybrid0.9 data
        if 'hybrid_09' in all_results[user_count]:
            hybrid_metrics = all_results[user_count]['hybrid_09']
            
            # Hybrid0.9 Latency row
            ws.cell(row=current_row, column=1, value='Hybrid0.9').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Latency').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            latency_cell = ws.cell(row=current_row, column=3, value=round(hybrid_metrics['avg_response'], 1))
            latency_cell.fill = latency_fill
            latency_cell.font = latency_font
            latency_cell.border = border
            latency_cell.alignment = center_align
            current_row += 1
            
            # Hybrid0.9 Throughput row
            ws.cell(row=current_row, column=1, value='Hybrid0.9').border = border
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2, value='Throughput').border = border
            ws.cell(row=current_row, column=2).alignment = left_align
            throughput_cell = ws.cell(row=current_row, column=3, value=round(hybrid_metrics['rps'], 2))
            throughput_cell.fill = throughput_fill
            throughput_cell.font = throughput_font
            throughput_cell.border = border
            throughput_cell.alignment = center_align
            current_row += 1
        
        # Add spacing between sections
        current_row += 2
    
    # Adjust column widths for better formatting
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    return wb


def main():
    """Main function"""
    print("=" * 70)
    print("EXCEL PERFORMANCE REPORT GENERATOR")
    print("=" * 70)
    print("\nScanning reports folders for all user counts...")
    print("-" * 70)
    
    # Scan all user counts
    all_results = scan_all_user_reports()
    
    if not all_results:
        print("\n❌ No data found!")
        print("   Make sure you have reports_RF*_Users*_Limit* folders with CSV files")
        return 1
    
    print("\n" + "-" * 70)
    print(f"✅ Found data for {len(all_results)} user counts")
    print(f"   User counts: {', '.join(sorted(all_results.keys(), key=lambda x: int(x)))}")
    
    # Get RF value
    rf_value = os.environ.get('PT_RF_VALUE', '3')
    
    # Generate Excel report
    print("\n📝 Generating Excel report...")
    wb = create_comparison_excel_report(all_results, rf_value=rf_value)
    
    # Save file
    output_file = f"../reports/multi_collection/multi_collection_performance_RF{rf_value}.xlsx"
    wb.save(output_file)
    
    print(f"✅ Created: {output_file}")
    print(f"   Location: {os.path.abspath(output_file)}")
    print("\n" + "=" * 70)
    print("EXCEL REPORT GENERATED!")
    print("=" * 70)
    print(f"\nOpen the report:")
    print(f"   {output_file}")
    print("=" * 70)
    
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())

